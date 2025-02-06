from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl

############################################################################################
## XLSX DATA RETRIEVAL PROCESS

path = "location of xlsx file on computer"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Count of the Rows with data on XLSX File
row_count = 0
for row in sheet_obj:
    if not all([cell.value == None for cell in row]):
        row_count += 1

# Count of the Columns on XLSX File
max_col = sheet_obj.max_column

# Creating a big list with all data values from XLSX File. 
# Going cell by cell, left to right, each row at a time.
xlsx_data_list = []
for i in range(2, row_count + 1):
    for j in range(1, max_col):
        cell_obj = sheet_obj.cell(row = i, column= j)
        xlsx_data_list.append(cell_obj.value)

# Shrinking the list to 7 elements at a time to represent a single users information. 
def list_creation(xlsx_data_list):
    input_list = [] 
    idx = 0
    for i in range(row_count -1): # 10
        for j in range(max_col- 1): # 7
            if idx == len(xlsx_data_list):
                break
            else:
                input_list.append(xlsx_data_list[idx])
                idx += 1
        data_input(input_list)

############################################################################################
## SELENIUM PROCESS

chrome_driver_path = "location of chromedriver.exe on computer"
driver = webdriver.Chrome(executable_path=chrome_driver_path)
driver.maximize_window()

# Open RPA website #
rpa_website = driver.get("http://www.rpachallenge.com/")

# Inputting the data onto the RPA Website 
def data_input(input_list):
    # "Getting" the start button on the RPA Website 
    start_button = driver.find_element_by_class_name('btn-large.uiColorButton')
    count = 0
    while count < 1:
        start_button.click()
        count += 1
    # "Getting" multiple form HTML elements on the RPA Website 
    first_name_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelFirstName"]')
    last_name_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelLastName"]')
    cmpny_name_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelCompanyName"]')
    role_in_cmpny_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelRole"]')
    address_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelAddress"]')
    email_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelEmail"]')
    phone_num_form = driver.find_element_by_xpath('//*[@ng-reflect-name="labelPhone"]')
    submit_button = driver.find_element_by_class_name('btn.uiColorButton')
    # Inputting the Data
    first_name_form.click()
    first_name_form.send_keys(input_list[0]) # 0 = First Name 
    last_name_form.click()
    last_name_form.send_keys(input_list[1]) # 1 = Last Name 
    cmpny_name_form.click()
    cmpny_name_form.send_keys(input_list[2]) # 2 = Company Name 
    role_in_cmpny_form.click()
    role_in_cmpny_form.send_keys(input_list[3]) # 3 = Role in Company 
    address_form.click()
    address_form.send_keys(input_list[4]) # 4 = Address
    email_form.click()
    email_form.send_keys(input_list[5]) # 5 = Email
    phone_num_form.click()
    phone_num_form.send_keys(input_list[6]) # 6 = Phone Number 
    submit_button.click()
    input_list.clear() # Clears the current data elements in the list
    return

list_creation(xlsx_data_list)

############################################################################################
